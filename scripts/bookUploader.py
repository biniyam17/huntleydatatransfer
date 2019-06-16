#dependencies
import openpyxl
import mysql.connector
import re

from config import port, database, user, password
from data_2018_fall import *
from data_utils import *

#Workbook Actions
# doc = openpyxl.load_workbook('../excel/CMC_Fall_2018_bookstore_list.xlsx', read_only = True, data_only = True)
doc = openpyxl.load_workbook('../excel/CMC_Spring_2019_bookstore_list.xlsx', read_only = True, data_only = True)
print (doc.sheetnames)
doc = doc.active
start_index = 2
startpoint = 'F' + str(start_index)
endpoint = 'M' + str(10)
sheet = doc[startpoint:endpoint]

#Database Actions
database = mysql.connector.connect(host='localhost',port=port,db=database, user=user, passwd=password)
cursor=database.cursor()

# verify_element_exists(sheet, dept_dict_18f,0)
# verify_element_exists(sheet, prof_dict_18f,2)
# insert_departments(sheet, cursor, database, dept_dict_18f)
# insert_professors(sheet, cursor, database, prof_dict_18f)
# insert_courses(sheet, cursor, database, 2, 22, 23, prof_dict_18f)
# insert_books(sheet, cursor, database, books_dict_18f)
# insert_course_books(sheet, cursor, database, books_dict_18f, course_code_to_id_map_18f)
# update_courses(sheet, cursor, database, 0, dept_dict_18f)
# write_google_books_api_responses(sheet, cursor, database,start_index)

# insert_departments_v2(sheet, cursor, database, 0)
# insert_professors_v2(sheet, cursor, database, 0)
# insert_courses_v2(sheet, cursor, database, 0, 3)
# write_google_books_api_responses_v2(sheet, cursor, database, start_index)
# insert_course_books_v2(sheet, cursor, database, 3, start_index)
insert_huntley_data(sheet, cursor, database, start_index)
